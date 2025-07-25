import tkinter as tk
from tkinter import messagebox
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

class SimpleFeedbackApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Simple Feedback Form")
        self.root.geometry("500x400")

        # Variables for inputs
        self.clarity_var = tk.IntVar(value=0)
        self.content_var = tk.IntVar(value=0)

        # Labels and radio buttons for ratings
        tk.Label(root, text="Rate Teacher's Clarity (1-5):").pack(pady=5)
        for i in range(1, 6):
            tk.Radiobutton(root, text=str(i), variable=self.clarity_var, value=i).pack(anchor='w')

        tk.Label(root, text="Rate Course Content (1-5):").pack(pady=5)
        for i in range(1, 6):
            tk.Radiobutton(root, text=str(i), variable=self.content_var, value=i).pack(anchor='w')

        # Comments box
        tk.Label(root, text="Additional Comments:").pack(pady=5)
        self.comments_text = tk.Text(root, height=5, width=40)
        self.comments_text.pack()

        # Submit button
        submit_btn = tk.Button(root, text="Submit Feedback", command=self.submit_feedback)
        submit_btn.pack(pady=20)

    def submit_feedback(self):
        if self.clarity_var.get() == 0 or self.content_var.get() == 0:
            messagebox.showwarning("Incomplete Form", "Please provide ratings for both questions.")
            return

        data = {
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Clarity Rating": self.clarity_var.get(),
            "Content Rating": self.content_var.get(),
            "Comments": self.comments_text.get("1.0", tk.END).strip()
        }

        excel_file = "feedback.xlsx"
        if not os.path.isfile(excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Feedback"
            headers = list(data.keys())
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            wb.save(excel_file)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append(list(data.values()))
        wb.save(excel_file)

        messagebox.showinfo("Success", "Feedback saved to Excel!")
        self.clear_form()

    def clear_form(self):
        self.clarity_var.set(0)
        self.content_var.set(0)
        self.comments_text.delete("1.0", tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = SimpleFeedbackApp(root)
    root.mainloop()
